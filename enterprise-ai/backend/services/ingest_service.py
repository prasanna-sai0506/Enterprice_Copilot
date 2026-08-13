"""Document ingest for RAG.

Chunking strategy
-----------------
1. Extract text page-by-page (PDF) or section-by-section (DOCX/XLSX/TXT).
2. Split on natural boundaries first (headings, paragraphs, sentences).
3. Target ~700 characters per chunk (~150–180 words) so embeddings stay focused.
4. Overlap 120 characters so a sentence that straddles a boundary is not lost.
5. Drop empty / tiny fragments (<40 chars).
6. Each chunk is stored with tenant_id so another company can never retrieve it.
"""

import io
import re
import uuid
from pathlib import Path

from langchain_text_splitters import RecursiveCharacterTextSplitter

from config import settings
from services.rag_service import add_chunks

SEPARATORS = ["\n## ", "\n# ", "\n\n", "\n", ". ", " ", ""]


def detect_type(filename: str) -> str:
    ext = Path(filename).suffix.lower().lstrip(".")
    mapping = {"pdf": "pdf", "docx": "docx", "xlsx": "xlsx", "txt": "txt", "md": "txt"}
    if ext not in mapping:
        raise ValueError(f"Unsupported file type: {ext}")
    return mapping[ext]


def extract_pages(file_bytes: bytes, filename: str, file_type: str) -> list[tuple[int, str]]:
    pages = []
    if file_type == "txt":
        pages.append((1, file_bytes.decode("utf-8", errors="ignore")))
    elif file_type == "pdf":
        import fitz

        doc = fitz.open(stream=file_bytes, filetype="pdf")
        for i, page in enumerate(doc, 1):
            pages.append((i, page.get_text() or ""))
    elif file_type == "docx":
        from docx import Document as Docx

        d = Docx(io.BytesIO(file_bytes))
        text = "\n".join(p.text for p in d.paragraphs)
        pages.append((1, text))
    elif file_type == "xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        for si, sheet in enumerate(wb.worksheets, 1):
            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append(" | ".join("" if c is None else str(c) for c in row))
            pages.append((si, f"Sheet {sheet.title}\n" + "\n".join(rows)))
    return pages


def _heading(text: str) -> str:
    line = (text or "").strip().splitlines()[0] if text else ""
    line = re.sub(r"^#+\s*", "", line).strip()
    return line[:80] if line else ""


def chunk_pages(pages: list[tuple[int, str]]) -> list[dict]:
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=700,
        chunk_overlap=120,
        separators=SEPARATORS,
        length_function=len,
    )
    out = []
    idx = 0
    for page_no, text in pages:
        for piece in splitter.split_text(text or ""):
            clean = piece.strip()
            if len(clean) < 40:
                continue
            out.append(
                {
                    "text": clean,
                    "page_number": page_no,
                    "chunk_index": idx,
                    "heading": _heading(clean),
                    "char_count": len(clean),
                }
            )
            idx += 1
    return out


def ingest_document(
    file_bytes: bytes,
    filename: str,
    tenant_id: str,
    namespace: str,
    dept: str,
    access_level: str,
    doc_id: str | None = None,
) -> tuple[str, int]:
    file_type = detect_type(filename)
    doc_id = doc_id or str(uuid.uuid4())
    pages = extract_pages(file_bytes, filename, file_type)
    pieces = chunk_pages(pages)
    ids, docs, metas = [], [], []
    # Company-wide default: general docs are visible to every employee of this tenant.
    level = access_level or "general"
    for p in pieces:
        ids.append(f"{doc_id}_{p['chunk_index']}")
        docs.append(p["text"])
        metas.append(
            {
                "tenant_id": tenant_id,
                "doc_id": doc_id,
                "filename": filename,
                "page_number": p["page_number"],
                "department": dept or "general",
                "access_level": level,
                "chunk_index": p["chunk_index"],
                "heading": p["heading"],
                "shared_with": "all_employees" if level == "general" else level,
            }
        )
    if docs:
        add_chunks(namespace, ids, docs, metas)
    dest_dir = Path(settings.UPLOAD_DIR) / tenant_id
    dest_dir.mkdir(parents=True, exist_ok=True)
    (dest_dir / f"{doc_id}_{filename}").write_bytes(file_bytes)
    return doc_id, len(docs)
